#!/usr/bin/env python3
"""Genera SQL de importación desde reportes de ventas legacy (XLSX)."""

from __future__ import annotations

import json
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
PUBLIC = ROOT / "public"
OUT_SQL = ROOT / "supabase" / "migrations" / "20260702130100_import_ventas_legacy_data.sql"
OUT_JSON = PUBLIC / "data" / "ventas-legacy-bundle.json"

FILES = [
    # 2024
    ("Setiembre2024.xlsx", "2024-09"),
    ("Octubre2024.xlsx", "2024-10"),
    ("Nov2024.xlsx", "2024-11"),
    ("Dic2024.xlsx", "2024-12"),
    # 2025
    ("Enero2025.xlsx", "2025-01"),
    ("Febrero25.xlsx", "2025-02"),
    ("Marzo.xlsx", "2025-03"),
    ("Abril.xlsx", "2025-04"),
    ("Mayo.xlsx", "2025-05"),
    ("Junio.xlsx", "2025-06"),
    ("Julio.xlsx", "2025-07"),
    ("Agosto.xlsx", "2025-08"),
    ("Setiembre.xlsx", "2025-09"),
    ("Octubre.xlsx", "2025-10"),
    ("Noviembre25.xlsx", "2025-11"),
    ("Diciembre.xlsx", "2025-12"),
    # 2026
    ("Enero.xlsx", "2026-01"),
    ("Febrero.xlsx", "2026-02"),
    ("Reporte_de_Ventas_2026070253857366.xlsx", "2026-03"),
    ("Abril2025.xlsx", "2026-04"),
    ("Reporte_de_Ventas_2026070253736502.xlsx", "2026-05"),
    ("Reporte_de_Ventas_2026070253617363.xlsx", "2026-06"),
]

DOC_TYPES = {"FACTURA", "BOLETA DE VENTA", "NOTA DE CREDITO", "ORDEN"}


def sql_literal(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


def sql_number(value: Decimal | float | int | None) -> str:
    if value is None:
        return "NULL"
    return str(value)


def is_red(cell) -> bool:
    if not cell or not cell.font or not cell.font.color:
        return False
    color = cell.font.color
    if color.type == "rgb" and color.rgb:
        return "FF0000" in str(color.rgb).upper()
    return False


def to_decimal(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except Exception:
        return None


def parse_fecha_hora(value) -> tuple[str | None, str | None]:
    if isinstance(value, datetime):
        return value.date().isoformat(), value.time().strftime("%H:%M:%S")
    if isinstance(value, date):
        return value.isoformat(), None
    if isinstance(value, str) and "/" in value:
        parts = value.strip().split("/")
        if len(parts) == 3:
            d, m, y = parts
            return f"{y}-{m.zfill(2)}-{d.zfill(2)}", None
    return None, None


def map_tipo(documento: str) -> str:
    doc = documento.strip().upper()
    if doc == "BOLETA DE VENTA":
        return "boleta"
    if doc == "NOTA DE CREDITO":
        return "nota_credito"
    return "factura"


def map_moneda(value: str | None) -> str:
    text = (value or "").strip().upper()
    if text in {"DOLARES", "USD", "US$"}:
        return "USD"
    return "PEN"


def format_codigo(serie: str, numero: int) -> str:
    return f"{serie.strip().upper()}-{numero:05d}"


def clean_ruc(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    digits = "".join(ch for ch in text if ch.isdigit())
    if not digits:
        return text
    if len(digits) <= 8:
        return digits.zfill(8)
    return digits.zfill(11) if len(digits) < 11 else digits


def calc_totals(importe_soles: Decimal | None, total_doc: Decimal | None) -> tuple[Decimal, Decimal, Decimal]:
    total = importe_soles if importe_soles is not None else (total_doc or Decimal("0"))
    if total == 0:
        return Decimal("0"), Decimal("0"), Decimal("0")
    subtotal = (total / Decimal("1.18")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    igv = (total - subtotal).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return subtotal, igv, total


def find_header_row(ws) -> int:
    for index, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and row[0] == "FECFAC":
            return index
    raise ValueError("No se encontró fila de encabezados FECFAC")


def parse_workbook(path: Path, periodo_mes: str, fuente: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_row = find_header_row(ws)
    rows: list[dict] = []

    for excel_row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        documento_cell = excel_row[2]
        documento = documento_cell.value
        if documento is None:
            break
        documento = str(documento).strip()
        if documento.upper().startswith("TOTAL"):
            break
        if documento not in DOC_TYPES:
            continue

        fecha, hora = parse_fecha_hora(excel_row[0].value)
        if not fecha:
            continue
        fecha_venc, _ = parse_fecha_hora(excel_row[1].value)
        if not fecha_venc:
            fecha_venc = fecha

        serie = str(excel_row[3].value or "").strip()
        numero = int(float(excel_row[4].value or 0))
        codigo = format_codigo(serie, numero)
        total_doc = to_decimal(excel_row[10].value)
        importe_soles = to_decimal(excel_row[12].value)
        anulada = any(is_red(cell) for cell in excel_row) or (
            total_doc is not None and total_doc > 0 and (importe_soles is None or importe_soles == 0)
        )
        subtotal, igv, total = calc_totals(importe_soles, total_doc)
        forma_pago = str(excel_row[14].value).strip() if excel_row[14].value not in (None, "", "-") else None
        cuenta = str(excel_row[15].value).strip() if excel_row[15].value not in (None, "", "-") else None
        numero_operacion = str(excel_row[16].value).strip() if excel_row[16].value not in (None, "", "-") else None
        tipo_cambio = to_decimal(excel_row[11].value)
        vendedor = str(excel_row[7].value).strip() if excel_row[7].value not in (None, "", "-") else None
        usuario = str(excel_row[8].value).strip() if excel_row[8].value not in (None, "", "-") else None
        cliente = str(excel_row[6].value or "").strip() or "Cliente"
        moneda_origen = map_moneda(str(excel_row[9].value) if excel_row[9].value else None)

        notas_parts = [f"Fuente: {fuente}", f"Periodo: {periodo_mes}"]
        if documento == "ORDEN":
            notas_parts.append("Tipo legacy: Orden de venta")
        if forma_pago:
            notas_parts.append(f"Forma de pago: {forma_pago}")
        if cuenta:
            notas_parts.append(f"Cuenta: {cuenta}")
        if numero_operacion:
            notas_parts.append(f"Operación: {numero_operacion}")
        if usuario:
            notas_parts.append(f"Usuario: {usuario}")
        if tipo_cambio is not None:
            notas_parts.append(f"T.C: {tipo_cambio}")
        if moneda_origen != "PEN":
            notas_parts.append(f"Moneda origen: {moneda_origen}")
        if anulada:
            notas_parts.append("Estado comercial: Anulada")

        row_data = {
            "codigo_comprobante": codigo,
            "numero": codigo,
            "fecha": fecha,
            "fecha_vencimiento": fecha_venc,
            "hora_emision": hora,
            "tipo_comprobante": map_tipo(documento),
            "serie": serie,
            "numero_correlativo": numero,
            "cliente_ruc": clean_ruc(excel_row[5].value),
            "cliente_nombre": cliente,
            "vendedor_nombre": vendedor,
            "usuario_emision": usuario,
            "moneda": "PEN",
            "moneda_origen": moneda_origen,
            "subtotal": float(subtotal),
            "igv": float(igv),
            "total": float(total),
            "tipo_cambio": float(tipo_cambio) if tipo_cambio is not None else None,
            "forma_pago": forma_pago,
            "cuenta_cobro": cuenta,
            "numero_operacion": numero_operacion,
            "periodo_mes": periodo_mes,
            "fuente_archivo": fuente,
            "notas": " · ".join(notas_parts),
            "estado": "anulada" if anulada else "confirmada",
            "estado_sunat": "aceptado",
            "tiene_cdr": not anulada,
            "anulada": anulada,
        }
        rows.append(row_data)

    return {
        "archivo": fuente,
        "periodo_mes": periodo_mes,
        "rows": rows,
        "total": len(rows),
        "anulados": sum(1 for row in rows if row["anulada"]),
    }


def main() -> None:
    bundles: list[dict] = []
    all_rows: list[dict] = []

    for filename, periodo in FILES:
        path = PUBLIC / filename
        if not path.exists():
            raise FileNotFoundError(path)
        bundle = parse_workbook(path, periodo, filename)
        bundles.append(bundle)
        all_rows.extend(bundle["rows"])
        print(f"{filename}: {bundle['total']} filas, {bundle['anulados']} anuladas")

    lines = [
        "-- Datos generados por scripts/import-ventas-xlsx.py",
        "TRUNCATE TABLE public.ventas_legacy_import;",
        "",
    ]

    for row in all_rows:
        lines.append(
            "INSERT INTO public.ventas_legacy_import ("
            "codigo_comprobante, numero, fecha, fecha_vencimiento, hora_emision, tipo_comprobante, serie, numero_correlativo, "
            "cliente_ruc, cliente_nombre, vendedor_nombre, usuario_emision, moneda, moneda_origen, subtotal, igv, total, "
            "tipo_cambio, forma_pago, cuenta_cobro, numero_operacion, periodo_mes, fuente_archivo, notas, estado, estado_sunat, tiene_cdr"
            ") VALUES ("
            f"{sql_literal(row['codigo_comprobante'])}, {sql_literal(row['numero'])}, {sql_literal(row['fecha'])}, "
            f"{sql_literal(row['fecha_vencimiento'])}, {sql_literal(row['hora_emision'])}, {sql_literal(row['tipo_comprobante'])}, "
            f"{sql_literal(row['serie'])}, {row['numero_correlativo']}, {sql_literal(row['cliente_ruc'])}, "
            f"{sql_literal(row['cliente_nombre'])}, {sql_literal(row['vendedor_nombre'])}, {sql_literal(row['usuario_emision'])}, "
            f"{sql_literal(row['moneda'])}, {sql_literal(row['moneda_origen'])}, "
            f"{sql_number(row['subtotal'])}, {sql_number(row['igv'])}, {sql_number(row['total'])}, "
            f"{sql_number(row['tipo_cambio'])}, {sql_literal(row['forma_pago'])}, {sql_literal(row['cuenta_cobro'])}, "
            f"{sql_literal(row['numero_operacion'])}, {sql_literal(row['periodo_mes'])}, {sql_literal(row['fuente_archivo'])}, "
            f"{sql_literal(row['notas'])}, {sql_literal(row['estado'])}, {sql_literal(row['estado_sunat'])}, "
            f"{'true' if row['tiene_cdr'] else 'false'}"
            ");"
        )

    OUT_SQL.write_text("\n".join(lines) + "\n", encoding="utf-8")

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(
        json.dumps(
            {
                "generatedAt": datetime.now().isoformat(),
                "files": bundles,
                "rows": all_rows,
                "summary": {
                    "total": len(all_rows),
                    "anulados": sum(1 for row in all_rows if row["anulada"]),
                },
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    print(f"Wrote {OUT_SQL} ({len(all_rows)} rows)")
    print(f"Wrote {OUT_JSON}")


if __name__ == "__main__":
    main()
