#!/usr/bin/env python3
"""Genera SQL y bundle JSON desde reportes de guías de remisión remitente (XLSX)."""

from __future__ import annotations

import json
import re
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
PUBLIC = ROOT / "public"
OUT_SQL = ROOT / "supabase" / "migrations" / "20260702160100_import_guias_legacy_data.sql"
OUT_JSON = PUBLIC / "data" / "guias-legacy-bundle.json"

FILES = [
    ("Guias Enero 2026.xlsx", "2026-01"),
    ("Guias Febrero 2026.xlsx", "2026-02"),
    ("Guias Marzo 2026.xlsx", "2026-03"),
    ("Guias Abril 2026.xlsx", "2026-04"),
    ("Guias Mayo 2026.xlsx", "2026-05"),
    ("Guias Junio.xlsx", "2026-06"),
]

HEADER_MARKERS = {"fecha emision", "serie", "numero"}


def sql_literal(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


def sql_number(value: Decimal | float | int | None) -> str:
    if value is None:
        return "NULL"
    return str(value)


def to_decimal(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except Exception:
        return None


def parse_fecha(value) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def format_codigo_guia(serie: str, numero: int) -> str:
    return f"{serie.strip().upper()}-{int(numero):05d}"


def normalize_comprobante(raw: str | None) -> str | None:
    if not raw:
        return None
    match = re.search(r"([A-Z]\d{3})-(\d+)", raw.upper())
    if not match:
        return None
    return f"{match.group(1)}-{int(match.group(2)):05d}"


def map_motivo(value: str | None) -> str:
    if not value:
        return "otros"
    return value.strip().lower().replace(" ", "_")


def map_estado(motivo: str) -> str:
    if motivo in {"venta", "consignacion"}:
        return "en_transito"
    return "emitida"


def find_header_row(ws) -> int:
    for row_idx in range(1, min(20, ws.max_row + 1)):
        labels = {
            str(ws.cell(row_idx, col).value or "").strip().lower()
            for col in range(1, min(6, ws.max_column + 1))
        }
        if HEADER_MARKERS.issubset(labels):
            return row_idx
    return 7


def cell_str(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_workbook(path: Path, periodo_mes: str, fuente: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_row = find_header_row(ws)

    reporte_desde = None
    reporte_hasta = None
    for row_idx in range(1, header_row):
        label = str(ws.cell(row_idx, 1).value or "").strip().upper()
        if label.startswith("DESDE"):
            reporte_desde = parse_fecha(ws.cell(row_idx, 2).value)
        if "HASTA" in label:
            reporte_hasta = parse_fecha(ws.cell(row_idx, 18).value) or parse_fecha(ws.cell(row_idx, 2).value)

    grouped: dict[str, dict] = {}

    for row_idx in range(header_row + 1, ws.max_row + 1):
        serie = cell_str(ws.cell(row_idx, 2).value)
        numero_raw = ws.cell(row_idx, 3).value
        if not serie or numero_raw in (None, ""):
            continue

        try:
            numero = int(numero_raw)
        except (TypeError, ValueError):
            continue

        codigo = format_codigo_guia(serie, numero)
        fecha_emision = parse_fecha(ws.cell(row_idx, 1).value)
        fecha_traslado = parse_fecha(ws.cell(row_idx, 6).value) or fecha_emision
        if not fecha_emision:
            continue

        motivo = map_motivo(cell_str(ws.cell(row_idx, 5).value))
        observacion = cell_str(ws.cell(row_idx, 26).value)
        comprobante = normalize_comprobante(observacion)

        if codigo not in grouped:
            grouped[codigo] = {
                "codigo_guia": codigo,
                "numero": codigo,
                "serie": serie.upper(),
                "numero_correlativo": numero,
                "fecha_emision": fecha_emision,
                "fecha_traslado": fecha_traslado or fecha_emision,
                "motivo_traslado": motivo,
                "sucursal": cell_str(ws.cell(row_idx, 4).value),
                "destinatario_ruc": cell_str(ws.cell(row_idx, 7).value),
                "destinatario_nombre": cell_str(ws.cell(row_idx, 8).value) or "Sin destinatario",
                "direccion_partida": cell_str(ws.cell(row_idx, 9).value),
                "direccion_destino": cell_str(ws.cell(row_idx, 10).value),
                "conductor_ruc": cell_str(ws.cell(row_idx, 18).value),
                "conductor_nombre": cell_str(ws.cell(row_idx, 19).value),
                "licencia": cell_str(ws.cell(row_idx, 20).value),
                "placa": cell_str(ws.cell(row_idx, 21).value),
                "peso_total": float(to_decimal(ws.cell(row_idx, 16).value) or 0),
                "bultos": int(ws.cell(row_idx, 17).value or 0) if ws.cell(row_idx, 17).value else None,
                "observacion": observacion,
                "comprobante_relacionado": comprobante,
                "estado": map_estado(motivo),
                "estado_sunat": "aceptado",
                "periodo_mes": periodo_mes,
                "fuente_archivo": fuente,
                "notas": f"Fuente: {fuente} · Periodo: {periodo_mes}"
                + (f" · Comprobante: {comprobante}" if comprobante else ""),
                "items": [],
            }

        grouped[codigo]["items"].append(
            {
                "codigo": cell_str(ws.cell(row_idx, 11).value),
                "descripcion": cell_str(ws.cell(row_idx, 12).value) or "Ítem sin descripción",
                "cantidad": float(to_decimal(ws.cell(row_idx, 13).value) or 1),
                "unidad": cell_str(ws.cell(row_idx, 14).value) or "UND",
                "peso_unitario": float(to_decimal(ws.cell(row_idx, 15).value) or 0) or None,
            }
        )

    guias = list(grouped.values())
    total_items = sum(len(g["items"]) for g in guias)

    return {
        "archivo": fuente,
        "periodo_mes": periodo_mes,
        "reporte_desde": reporte_desde,
        "reporte_hasta": reporte_hasta,
        "guias": guias,
        "total_guias": len(guias),
        "total_items": total_items,
    }


def main() -> None:
    bundles: list[dict] = []
    all_guias: list[dict] = []

    for filename, periodo in FILES:
        path = PUBLIC / filename
        if not path.exists():
            raise FileNotFoundError(path)
        bundle = parse_workbook(path, periodo, filename)
        bundles.append(bundle)
        all_guias.extend(bundle["guias"])
        print(f"{filename}: {bundle['total_guias']} guías, {bundle['total_items']} ítems")

    lines = [
        "-- Datos generados por scripts/import-guias-xlsx.py",
        "TRUNCATE TABLE public.guias_legacy_import, public.guia_legacy_import_items RESTART IDENTITY CASCADE;",
        "",
    ]

    for guia in all_guias:
        lines.append(
            "INSERT INTO public.guias_legacy_import ("
            "codigo_guia, numero, serie, numero_correlativo, fecha_emision, fecha_traslado, "
            "motivo_traslado, sucursal, destinatario_ruc, destinatario_nombre, direccion_partida, "
            "direccion_destino, conductor_ruc, conductor_nombre, licencia, placa, peso_total, bultos, "
            "observacion, comprobante_relacionado, estado, estado_sunat, periodo_mes, fuente_archivo, notas"
            ") VALUES ("
            f"{sql_literal(guia['codigo_guia'])}, {sql_literal(guia['numero'])}, {sql_literal(guia['serie'])}, "
            f"{guia['numero_correlativo']}, {sql_literal(guia['fecha_emision'])}, {sql_literal(guia['fecha_traslado'])}, "
            f"{sql_literal(guia['motivo_traslado'])}, {sql_literal(guia['sucursal'])}, {sql_literal(guia['destinatario_ruc'])}, "
            f"{sql_literal(guia['destinatario_nombre'])}, {sql_literal(guia['direccion_partida'])}, "
            f"{sql_literal(guia['direccion_destino'])}, {sql_literal(guia['conductor_ruc'])}, "
            f"{sql_literal(guia['conductor_nombre'])}, {sql_literal(guia['licencia'])}, {sql_literal(guia['placa'])}, "
            f"{sql_number(guia['peso_total'])}, {sql_number(guia['bultos'])}, {sql_literal(guia['observacion'])}, "
            f"{sql_literal(guia['comprobante_relacionado'])}, {sql_literal(guia['estado'])}, "
            f"{sql_literal(guia['estado_sunat'])}, {sql_literal(guia['periodo_mes'])}, "
            f"{sql_literal(guia['fuente_archivo'])}, {sql_literal(guia['notas'])}"
            ");"
        )

        for item in guia["items"]:
            lines.append(
                "INSERT INTO public.guia_legacy_import_items ("
                "codigo_guia, codigo, descripcion, cantidad, unidad, peso_unitario"
                ") VALUES ("
                f"{sql_literal(guia['codigo_guia'])}, {sql_literal(item['codigo'])}, "
                f"{sql_literal(item['descripcion'])}, {sql_number(item['cantidad'])}, "
                f"{sql_literal(item['unidad'])}, {sql_number(item.get('peso_unitario'))}"
                ");"
            )

    OUT_SQL.write_text("\n".join(lines) + "\n", encoding="utf-8")

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(
        json.dumps(
            {
                "generatedAt": datetime.now().isoformat(),
                "files": bundles,
                "guias": all_guias,
                "summary": {
                    "total_guias": len(all_guias),
                    "total_items": sum(len(g["items"]) for g in all_guias),
                },
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    print(f"Wrote {OUT_SQL} ({len(all_guias)} guías)")
    print(f"Wrote {OUT_JSON}")


if __name__ == "__main__":
    main()
