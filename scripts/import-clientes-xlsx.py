#!/usr/bin/env python3
"""Genera SQL de importación desde Reporte_Persona (Rapifac) o formato legacy."""

from __future__ import annotations

import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "supabase" / "migrations" / "clientes_2026-08-19.xlsx"
OUT_SQL = ROOT / "supabase" / "migrations" / "20260819120000_import_clientes_persona_data.sql"

DIST_KEYWORDS = (
    "COPY",
    "SISTEMAS",
    "SOLUCIONES",
    "SUMINISTROS",
    "TECNOLOGIAS",
    "IMPORT",
    "DISTRIBUID",
    "COMPU",
    "COPIER",
    "DIGITAL",
)

PLACEHOLDER_EMAILS = {"no-send@rapifac.com", "no-send@rapifac.com.br"}

UBIGEO_FALLBACK: dict[str, str] = {
    "010100": "Amazonas, Chachapoyas, Chachapoyas",
    "020101": "Áncash, Huaraz, Huaraz",
    "020105": "Áncash, Huaraz, Independencia",
    "020801": "Áncash, Santa, Santa",
    "021101": "Áncash, Huarmey, Huarmey",
    "021801": "Áncash, Santa, Chimbote",
    "021809": "Áncash, Santa, Nuevo Chimbote",
    "040101": "Arequipa, Arequipa, Arequipa",
    "040104": "Arequipa, Arequipa, Cerro Colorado",
    "050101": "Ayacucho, Huamanga, Ayacucho",
    "060101": "Cajamarca, Cajamarca, Cajamarca",
    "060108": "Cajamarca, Cajamarca, Los Baños del Inca",
    "070101": "Callao, Callao, Callao",
    "070102": "Callao, Callao, Bellavista",
    "070103": "Callao, Callao, Carmen de la Legua Reynoso",
    "070106": "Callao, Callao, Ventanilla",
    "080101": "Cusco, Cusco, Cusco",
    "080108": "Cusco, Cusco, Wanchaq",
    "090101": "Huancavelica, Huancavelica, Huancavelica",
    "100101": "Huánuco, Huánuco, Huánuco",
    "100102": "Huánuco, Huánuco, Amarilis",
    "110101": "Ica, Ica, Ica",
    "110201": "Ica, Chincha, Chincha Alta",
    "110508": "Ica, Pisco, Túpac Amaru Inca",
    "120101": "Junín, Huancayo, Huancayo",
    "120302": "Junín, Chupaca, Chupaca",
    "130101": "La Libertad, Trujillo, Trujillo",
    "140101": "Lambayeque, Chiclayo, Chiclayo",
    "150101": "Lima, Lima, Lima",
    "150103": "Lima, Lima, Ate",
    "150105": "Lima, Lima, Breña",
    "150112": "Lima, Lima, Lurigancho",
    "150114": "Lima, Lima, La Molina",
    "150115": "Lima, Lima, Los Olivos",
    "150116": "Lima, Lima, San Juan de Lurigancho",
    "150117": "Lima, Lima, San Isidro",
    "150130": "Lima, Lima, Villa El Salvador",
    "150131": "Lima, Lima, Villa María del Triunfo",
    "150140": "Lima, Lima, Surquillo",
    "150142": "Lima, Lima, Santiago de Surco",
    "200101": "Piura, Piura, Piura",
    "250101": "Ucayali, Coronel Portillo, Pucallpa",
}


def sql_literal(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


def clean_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text in {"-", "—"}:
        return None
    return text


def clean_email(value) -> str | None:
    email = clean_text(value)
    if not email:
        return None
    lowered = email.lower()
    if lowered in PLACEHOLDER_EMAILS or lowered.startswith("http"):
        return None
    return email


def clean_ruc(value) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    digits = re.sub(r"\D", "", text)
    if not digits:
        return None
    if len(digits) <= 8:
        return digits.zfill(8)
    return digits.zfill(11) if len(digits) < 11 else digits


def resolve_ciudad_from_direccion(direccion: str | None) -> str | None:
    if not direccion:
        return None
    parts = [part.strip() for part in str(direccion).replace(" - ", " ").split() if part.strip()]
    raw = direccion.replace(" - ", " - ")
    segments = [part.strip() for part in raw.split(" - ") if part.strip() and part.strip() != "-"]
    if len(segments) >= 3:
        dept, prov, dist = segments[-3], segments[-2], segments[-1]
        return f"{dept.title()}, {prov.title()}, {dist.title()}"
    if len(segments) >= 1:
        return segments[-1].title()
    return None


def resolve_ciudad(ciudad_raw, direccion: str | None) -> str | None:
    ciudad = clean_text(ciudad_raw)
    if ciudad and not ciudad.isdigit():
        return ciudad.title() if ciudad.isupper() else ciudad

    from_dir = resolve_ciudad_from_direccion(direccion)
    if from_dir:
        return from_dir

    if ciudad and ciudad in UBIGEO_FALLBACK:
        return UBIGEO_FALLBACK[ciudad]

    return ciudad


def normalize_upper(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    return (
        unicodedata.normalize("NFD", text)
        .encode("ascii", "ignore")
        .decode("ascii")
        .upper()
        .strip()
    )


def resolve_tipo_cliente(nombre: str, categoria: str | None, tipo_precio: str | None) -> str:
    cat = normalize_upper(categoria)
    precio = normalize_upper(tipo_precio)
    upper_name = nombre.upper()

    if "DISTRIBUID" in cat or "DISTRIBUID" in precio:
        return "Distribuidor"
    if "TECNIC" in cat:
        return "Técnico"
    if "MAYORIST" in cat or "MAYORIST" in precio:
        return "Mayorista"
    if any(keyword in upper_name for keyword in DIST_KEYWORDS):
        return "Distribuidor"
    return "Público"


def resolve_segmento(tipo_cliente: str, categoria: str | None) -> str:
    cat = normalize_upper(categoria)
    if "CORPORATIV" in cat or tipo_cliente == "Distribuidor":
        return "Corporativo"
    if tipo_cliente == "Técnico":
        return "PYME"
    if "MAYORIST" in cat:
        return "Mayorista"
    return "Otros"


def resolve_estado(excel_estado: str | None) -> str:
    estado = normalize_upper(excel_estado)
    if estado in {"HABILITADO", "ACTIVO"}:
        return "activo"
    if estado in {"DESHABILITADO", "INACTIVO"}:
        return "inactivo"
    if estado in {"PROSPECTO"}:
        return "prospecto"
    return "activo"


def is_rapifac_format(headers: tuple) -> bool:
    joined = " ".join(str(h or "") for h in headers).upper()
    return "NUMERO DE DOCUMENTO" in joined or "NÚMERO DE DOCUMENTO" in joined


def parse_rapifac_row(row: tuple) -> dict | None:
    ruc = clean_ruc(row[1])
    nombre = clean_text(row[2])
    if not ruc or not nombre:
        return None

    direccion = clean_text(row[3])
    referencia = clean_text(row[4])
    if referencia and direccion:
        direccion = f"{direccion} ({referencia})"
    elif referencia:
        direccion = referencia

    correo = clean_email(row[5]) or clean_email(row[6])
    telefono = clean_text(row[7])
    ubigeo = clean_text(row[8])
    estado_raw = clean_text(row[10]) if len(row) > 10 else None
    categoria = clean_text(row[14]) if len(row) > 14 else None
    tipo_precio = clean_text(row[13]) if len(row) > 13 else None
    observaciones = clean_text(row[22]) if len(row) > 22 else None

    tipo_cliente = resolve_tipo_cliente(nombre, categoria, tipo_precio)
    return {
        "razon_social": nombre,
        "ruc": ruc,
        "correo": correo,
        "telefono": telefono,
        "direccion": direccion,
        "ciudad": resolve_ciudad(ubigeo, direccion),
        "tipo_cliente": tipo_cliente,
        "estado": resolve_estado(estado_raw),
        "observaciones": observaciones,
        "segmento": resolve_segmento(tipo_cliente, categoria),
    }


def parse_legacy_row(row: tuple) -> dict | None:
    nombre = clean_text(row[0]) or ""
    ruc = clean_ruc(row[1])
    if not ruc:
        return None

    direccion = clean_text(row[4])
    tipo_cliente = resolve_tipo_cliente(nombre, clean_text(row[6]), None)
    return {
        "razon_social": nombre,
        "ruc": ruc,
        "correo": clean_email(row[2]),
        "telefono": clean_text(row[3]),
        "direccion": direccion,
        "ciudad": resolve_ciudad(row[5], direccion),
        "tipo_cliente": tipo_cliente,
        "estado": resolve_estado(clean_text(row[7])),
        "observaciones": clean_text(row[10]) if len(row) > 10 else None,
        "segmento": resolve_segmento(tipo_cliente, clean_text(row[6])),
    }


def row_to_record(headers: tuple, row: tuple) -> dict | None:
    if is_rapifac_format(headers):
        return parse_rapifac_row(row)
    return parse_legacy_row(row)


def main() -> None:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = wb["Hoja1"] if "Hoja1" in wb.sheetnames else wb.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = rows[0]
    data_rows = [row for row in rows[1:] if row and any(cell is not None for cell in row)]

    lines: list[str] = [
        f"-- Datos generados por scripts/import-clientes-xlsx.py desde {xlsx_path.name}",
        "TRUNCATE TABLE public.clientes_legacy_import;",
        "",
    ]

    inserted = 0
    for row in data_rows:
        record = row_to_record(headers, row)
        if not record:
            continue
        inserted += 1
        lines.append(
            "INSERT INTO public.clientes_legacy_import "
            "(razon_social, ruc, correo, telefono, direccion, ciudad, tipo_cliente, estado, observaciones, segmento) VALUES ("
            f"{sql_literal(record['razon_social'])}, {sql_literal(record['ruc'])}, {sql_literal(record['correo'])}, "
            f"{sql_literal(record['telefono'])}, {sql_literal(record['direccion'])}, {sql_literal(record['ciudad'])}, "
            f"{sql_literal(record['tipo_cliente'])}, {sql_literal(record['estado'])}, "
            f"{sql_literal(record['observaciones'])}, {sql_literal(record['segmento'])}"
            ");"
        )

    OUT_SQL.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Generated {inserted} rows -> {OUT_SQL}")


if __name__ == "__main__":
    main()
