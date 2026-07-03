#!/usr/bin/env python3
"""Genera SQL de importación desde productos-inventario-2026-07-02.xlsx"""

from __future__ import annotations

import json
import re
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "public" / "productos-inventario-2026-07-02.xlsx"
OUT_SQL = ROOT / "supabase" / "migrations" / "20260702150100_import_productos_legacy_data.sql"
OUT_JSON = ROOT / "public" / "data" / "productos-legacy-bundle.json"

DEFAULT_ALMACEN = "Almacén Central"


def normalize_header(value: str | None) -> str:
    if not value:
        return ""
    text = str(value)
    replacements = {
        "ó": "o",
        "é": "e",
        "í": "i",
        "á": "a",
        "ú": "u",
        "ñ": "n",
        "Ó": "O",
        "É": "E",
        "Í": "I",
        "Á": "A",
        "Ú": "U",
        "Ñ": "N",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return text.strip()


def sql_literal(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


def sql_bool(value: bool) -> str:
    return "true" if value else "false"


def sql_number(value: Decimal | float | int | None) -> str:
    if value is None:
        return "NULL"
    return str(value)


def to_decimal(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except Exception:
        return None


def to_int(value) -> int:
    if value is None or value == "":
        return 0
    try:
        return max(0, int(Decimal(str(value))))
    except Exception:
        return 0


def resolve_category(raw: str | None) -> str:
    if not raw:
        return "General"
    text = str(raw).strip()
    if "," in text:
        return text.split(",")[0].strip() or "General"
    return text[:120]


def resolve_tipo(categoria: str, titulo: str) -> str:
    cat = categoria.lower()
    title = titulo.lower()
    if any(token in cat for token in ("software", "servicio")) or "licencia" in title:
        return "service"
    if "kit" in title:
        return "kit"
    return "product"


def resolve_unidad(tipo: str) -> str:
    if tipo == "service":
        return "servicio"
    if tipo == "kit":
        return "kit"
    return "und"


def resolve_precio_venta(publico, distribuidor, tecnico, mayorista) -> Decimal:
    for value in (publico, distribuidor, tecnico, mayorista):
        if value is not None and value > 0:
            return value
    return Decimal("0.00")


def build_descripcion(
    atributos: str | None,
    proveedores: str | None,
    url_imagen: str | None,
    mayorista: Decimal | None,
    tecnico: Decimal | None,
    distribuidor: Decimal | None,
    publico: Decimal | None,
) -> str | None:
    parts: list[str] = []
    if atributos:
        parts.append(str(atributos).strip())
    if proveedores:
        parts.append(f"Proveedor: {str(proveedores).strip()}")
    if url_imagen:
        parts.append(f"Imagen: {str(url_imagen).strip()}")

    tiers: list[str] = []
    if mayorista and mayorista > 0:
        tiers.append(f"Mayorista USD {mayorista}")
    if tecnico and tecnico > 0:
        tiers.append(f"Tecnico USD {tecnico}")
    if distribuidor and distribuidor > 0:
        tiers.append(f"Distribuidor USD {distribuidor}")
    if publico and publico > 0:
        tiers.append(f"Publico USD {publico}")
    if tiers:
        parts.append("Precios — " + " · ".join(tiers))

    if not parts:
        return None
    return " · ".join(parts)[:4000]


def resolve_sku(codigo: str, legacy_id: str, seen_codes: dict[str, str]) -> str:
    base = re.sub(r"\s+", " ", codigo.strip()) if codigo else ""
    if not base:
        return legacy_id[:80]
    if base not in seen_codes:
        seen_codes[base] = legacy_id
        return base[:80]
    if seen_codes[base] == legacy_id:
        return base[:80]
    suffix = legacy_id.split("-")[-1][:12]
    return f"{base}-{suffix}"[:80]


def parse_rows() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active
    headers = [normalize_header(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    seen_codes: dict[str, str] = {}
    rows: list[dict] = []

    for row_idx in range(2, ws.max_row + 1):
        raw = {headers[i]: ws.cell(row_idx, i + 1).value for i in range(len(headers))}
        legacy_id = str(raw.get("ID") or "").strip()
        titulo = str(raw.get("Titulo") or "").strip()
        if not legacy_id or not titulo:
            continue

        codigo = str(raw.get("Codigo") or "").strip()
        categoria_raw = str(raw.get("Categoria") or "General").strip()
        categoria = resolve_category(categoria_raw)
        tipo = resolve_tipo(categoria_raw, titulo)
        stock = 0 if tipo == "service" else to_int(raw.get("Stock"))
        costo = to_decimal(raw.get("Precio compra (USD)")) or Decimal("0.00")
        mayorista = to_decimal(raw.get("Precio Mayorista (USD)"))
        tecnico = to_decimal(raw.get("Precio Tecnico (USD)"))
        distribuidor = to_decimal(raw.get("Precio Distribuidor (USD)"))
        publico = to_decimal(raw.get("Precio Publico (USD)"))
        precio = resolve_precio_venta(publico, distribuidor, tecnico, mayorista)
        marca = str(raw.get("Marca") or "").strip() or None
        descripcion = build_descripcion(
            str(raw.get("Atributos") or "").strip() or None,
            str(raw.get("Proveedores") or "").strip() or None,
            str(raw.get("URL imagen") or "").strip() or None,
            mayorista,
            tecnico,
            distribuidor,
            publico,
        )

        rows.append(
            {
                "legacy_id": legacy_id,
                "sku": resolve_sku(codigo, legacy_id, seen_codes),
                "nombre": titulo[:250],
                "marca": marca,
                "categoria": categoria,
                "stock": stock,
                "costo": costo,
                "precio": precio,
                "precio_mayorista": mayorista,
                "precio_tecnico": tecnico,
                "precio_distribuidor": distribuidor,
                "moneda": str(raw.get("Moneda") or "USD").strip().upper() or "USD",
                "tipo": tipo,
                "unidad": resolve_unidad(tipo),
                "almacen": DEFAULT_ALMACEN,
                "afectacion_igv": "afecto",
                "descripcion": descripcion,
                "activo": True,
            }
        )

    return rows


def write_sql(rows: list[dict]) -> None:
    lines = [
        "-- Datos legacy de productos / inventario (generado automáticamente)",
        "",
        "TRUNCATE TABLE public.productos_legacy_import;",
        "",
    ]

    for row in rows:
        lines.append(
            "INSERT INTO public.productos_legacy_import ("
            "legacy_id, sku, nombre, marca, categoria, stock, costo, precio, "
            "precio_mayorista, precio_tecnico, precio_distribuidor, moneda, tipo, unidad, "
            "almacen, afectacion_igv, descripcion, activo"
            ") VALUES ("
            f"{sql_literal(row['legacy_id'])}, "
            f"{sql_literal(row['sku'])}, "
            f"{sql_literal(row['nombre'])}, "
            f"{sql_literal(row['marca'])}, "
            f"{sql_literal(row['categoria'])}, "
            f"{row['stock']}, "
            f"{sql_number(row['costo'])}, "
            f"{sql_number(row['precio'])}, "
            f"{sql_number(row['precio_mayorista'])}, "
            f"{sql_number(row['precio_tecnico'])}, "
            f"{sql_number(row['precio_distribuidor'])}, "
            f"{sql_literal(row['moneda'])}, "
            f"{sql_literal(row['tipo'])}, "
            f"{sql_literal(row['unidad'])}, "
            f"{sql_literal(row['almacen'])}, "
            f"{sql_literal(row['afectacion_igv'])}, "
            f"{sql_literal(row['descripcion'])}, "
            f"{sql_bool(row['activo'])}"
            ");"
        )

    OUT_SQL.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_json(rows: list[dict]) -> None:
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)

    def serialize(row: dict) -> dict:
        output: dict = {}
        for key, value in row.items():
            if isinstance(value, Decimal):
                output[key] = float(value)
            else:
                output[key] = value
        return output

    payload = {
        "source": str(XLSX.name),
        "total": len(rows),
        "with_stock": sum(1 for row in rows if row["stock"] > 0),
        "services": sum(1 for row in rows if row["tipo"] == "service"),
        "sample": [serialize(row) for row in rows[:5]],
    }
    OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    rows = parse_rows()
    write_sql(rows)
    write_json(rows)
    print(f"Productos procesados: {len(rows)}")
    print(f"SQL: {OUT_SQL}")
    print(f"JSON: {OUT_JSON}")


if __name__ == "__main__":
    main()
