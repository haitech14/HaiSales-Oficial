#!/usr/bin/env python3
"""Genera SQL y bundle JSON desde reportes Venta por ítem (XLSX)."""

from __future__ import annotations

import json
import re
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
PUBLIC = ROOT / "public"
MIGRATIONS = ROOT / "supabase" / "migrations"
OUT_SQL = ROOT / "supabase" / "migrations" / "20260702170100_import_venta_items_legacy_data.sql"
OUT_JSON = PUBLIC / "data" / "venta-items-legacy-bundle.json"

FILES = [
    ("Ventas_por_Item_202607026382245.xlsx", "2024-09"),
    ("Ventas_por_Item_2026070263755983.xlsx", "2024-10"),
    ("Ventas_por_Item_202607026375050.xlsx", "2024-11"),
    ("Ventas_por_Item_2026070263742874.xlsx", "2024-12"),
    ("Venta por item Enero 2025.xlsx", "2025-01"),
    ("Venta por item Febrero 2025.xlsx", "2025-02"),
    ("Venta por item Marzo 2025.xlsx", "2025-03"),
    ("Venta por item Abril 2025.xlsx", "2025-04"),
    ("Venta por item Mayo 2025.xlsx", "2025-05"),
    ("Venta pro item Junio 2025.xlsx", "2025-06"),
    ("Venta por item Julio 2025.xlsx", "2025-07"),
    ("Venta por item Agosto 2025.xlsx", "2025-08"),
    ("Venta por item Setiembre 2025.xlsx", "2025-09"),
    ("Venta por item Octubre 2025.xlsx", "2025-10"),
    ("Venta pro item Noviembre 2025.xlsx", "2025-11"),
    ("Venta pro item Diciembre 2025.xlsx", "2025-12"),
    ("Venta pro item Enero 2026.xlsx", "2026-01"),
    ("Venta pro item Febrero 2026.xlsx", "2026-02"),
    ("Venta por item Marzo 2026.xlsx", "2026-03"),
    ("Venta por item Abril 2026.xlsx", "2026-04"),
    ("Venta pro item Mayo 2026.xlsx", "2026-05"),
    ("Venta pro item Junio 2026.xlsx", "2026-06"),
]

DOC_TYPES = {"FACTURA", "BOLETA DE VENTA", "BOLETA", "NOTA DE CREDITO", "ORDEN"}

SCHEMA_PREAMBLE = """\
-- Asegura tabla si solo se ejecuta este archivo en el editor SQL
CREATE TABLE IF NOT EXISTS public.venta_legacy_import_items (
  id BIGSERIAL PRIMARY KEY,
  codigo_comprobante TEXT NOT NULL,
  numero TEXT NOT NULL,
  fecha DATE NOT NULL,
  tipo_comprobante TEXT NOT NULL,
  serie TEXT NOT NULL,
  numero_correlativo INTEGER NOT NULL,
  cliente_ruc TEXT,
  cliente_nombre TEXT NOT NULL,
  sucursal TEXT,
  categoria TEXT,
  tipo_producto TEXT,
  subtipo TEXT,
  codigo TEXT,
  descripcion TEXT NOT NULL,
  cantidad NUMERIC(14, 4) NOT NULL DEFAULT 0,
  unidad TEXT NOT NULL DEFAULT 'UNIDAD',
  precio_unitario NUMERIC(14, 4) NOT NULL DEFAULT 0,
  subtotal NUMERIC(14, 2) NOT NULL DEFAULT 0,
  periodo_mes TEXT NOT NULL,
  fuente_archivo TEXT NOT NULL,
  linea_orden INTEGER NOT NULL DEFAULT 0
);

CREATE INDEX IF NOT EXISTS venta_legacy_import_items_comprobante_idx
  ON public.venta_legacy_import_items (codigo_comprobante);

ALTER TABLE public.venta_legacy_import_items ENABLE ROW LEVEL SECURITY;

DROP POLICY IF EXISTS "Authenticated can read venta items legacy import" ON public.venta_legacy_import_items;
CREATE POLICY "Authenticated can read venta items legacy import" ON public.venta_legacy_import_items
  FOR SELECT TO authenticated USING (true);
"""


def sql_literal(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


def sql_number(value: Decimal | float | int | None) -> str:
    if value is None:
        return "NULL"
    return str(value)


def to_decimal(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
    except Exception:
        return None


def parse_fecha(value) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def clean_ruc(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    digits = "".join(ch for ch in text if ch.isdigit())
    if not digits:
        return text
    if len(digits) <= 8:
        return digits.zfill(8)
    return digits.zfill(11) if len(digits) < 11 else digits


def map_tipo(documento: str) -> str:
    doc = documento.strip().upper()
    if "BOLETA" in doc:
        return "boleta"
    if "NOTA" in doc:
        return "nota_credito"
    return "factura"


def format_codigo(serie: str, numero: int) -> str:
    return f"{serie.strip().upper()}-{int(numero):05d}"


def cell_str(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_header(value) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("ó", "o").replace("í", "i").replace("á", "a").replace("é", "e").replace("ú", "u")
    text = text.replace("ñ", "n")
    return re.sub(r"\s+", " ", text)


def find_header_row(ws) -> int:
    for row_idx in range(1, min(20, ws.max_row + 1)):
        first = normalize_header(ws.cell(row_idx, 1).value)
        second = normalize_header(ws.cell(row_idx, 2).value)
        if first == "fecha" and second == "documento":
            return row_idx
    raise ValueError("No se encontró fila de encabezados Fecha/Documento")


def parse_workbook(path: Path, periodo_mes: str, fuente: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_row = find_header_row(ws)
    items: list[dict] = []
    linea = 0

    for excel_row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        documento = cell_str(excel_row[1] if len(excel_row) > 1 else None)
        if not documento:
            break
        if documento.upper().startswith("TOTAL"):
            break
        if documento.upper() not in DOC_TYPES:
            continue

        fecha = parse_fecha(excel_row[0])
        if not fecha:
            continue

        serie = cell_str(excel_row[2]) or ""
        numero = int(float(excel_row[3] or 0))
        if not serie or numero <= 0:
            continue

        codigo_comprobante = format_codigo(serie, numero)
        descripcion = cell_str(excel_row[13]) or "Producto"
        codigo = cell_str(excel_row[12])
        cantidad = to_decimal(excel_row[15]) or Decimal("1")
        unidad = cell_str(excel_row[16]) or "UNIDAD"
        precio_unitario = to_decimal(excel_row[17]) or Decimal("0")
        subtotal = to_decimal(excel_row[20])
        if subtotal is None or subtotal == 0:
            subtotal = (cantidad * precio_unitario).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        linea += 1
        items.append(
            {
                "codigo_comprobante": codigo_comprobante,
                "numero": codigo_comprobante,
                "fecha": fecha,
                "tipo_comprobante": map_tipo(documento),
                "serie": serie.upper(),
                "numero_correlativo": numero,
                "cliente_ruc": clean_ruc(excel_row[5]),
                "cliente_nombre": cell_str(excel_row[6]) or "Cliente",
                "sucursal": cell_str(excel_row[4]),
                "categoria": cell_str(excel_row[9]),
                "tipo_producto": cell_str(excel_row[10]),
                "subtipo": cell_str(excel_row[11]),
                "codigo": codigo,
                "descripcion": descripcion[:500],
                "cantidad": float(cantidad),
                "unidad": unidad,
                "precio_unitario": float(precio_unitario),
                "subtotal": float(subtotal.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
                "periodo_mes": periodo_mes,
                "fuente_archivo": fuente,
                "linea_orden": linea,
            }
        )

    comprobantes = len({item["codigo_comprobante"] for item in items})
    return {
        "fuente": fuente,
        "periodo_mes": periodo_mes,
        "total_items": len(items),
        "total_comprobantes": comprobantes,
        "items": items,
    }


def resolve_xlsx(filename: str) -> Path:
    for base in (PUBLIC, MIGRATIONS):
        path = base / filename
        if path.exists():
            return path
    raise FileNotFoundError(filename)


def main() -> None:
    bundles: list[dict] = []
    all_items: list[dict] = []

    for filename, periodo in FILES:
        path = resolve_xlsx(filename)
        bundle = parse_workbook(path, periodo, filename)
        bundles.append(bundle)
        all_items.extend(bundle["items"])
        print(
            f"{filename}: {bundle['total_comprobantes']} comprobantes, "
            f"{bundle['total_items']} ítems"
        )

    lines = [
        "-- Datos generados por scripts/import-venta-items-xlsx.py",
        SCHEMA_PREAMBLE.rstrip(),
        "",
        "TRUNCATE TABLE public.venta_legacy_import_items;",
        "",
    ]

    for item in all_items:
        lines.append(
            "INSERT INTO public.venta_legacy_import_items ("
            "codigo_comprobante, numero, fecha, tipo_comprobante, serie, numero_correlativo, "
            "cliente_ruc, cliente_nombre, sucursal, categoria, tipo_producto, subtipo, codigo, "
            "descripcion, cantidad, unidad, precio_unitario, subtotal, periodo_mes, fuente_archivo, linea_orden"
            ") VALUES ("
            f"{sql_literal(item['codigo_comprobante'])}, {sql_literal(item['numero'])}, "
            f"{sql_literal(item['fecha'])}, {sql_literal(item['tipo_comprobante'])}, "
            f"{sql_literal(item['serie'])}, {item['numero_correlativo']}, "
            f"{sql_literal(item['cliente_ruc'])}, {sql_literal(item['cliente_nombre'])}, "
            f"{sql_literal(item['sucursal'])}, {sql_literal(item['categoria'])}, "
            f"{sql_literal(item['tipo_producto'])}, {sql_literal(item['subtipo'])}, "
            f"{sql_literal(item['codigo'])}, {sql_literal(item['descripcion'])}, "
            f"{sql_number(item['cantidad'])}, {sql_literal(item['unidad'])}, "
            f"{sql_number(item['precio_unitario'])}, {sql_number(item['subtotal'])}, "
            f"{sql_literal(item['periodo_mes'])}, {sql_literal(item['fuente_archivo'])}, "
            f"{item['linea_orden']}"
            ");"
        )

    OUT_SQL.write_text("\n".join(lines) + "\n", encoding="utf-8")

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(
        json.dumps(
            {
                "generatedAt": datetime.now().isoformat(),
                "files": bundles,
                "summary": {
                    "total_items": len(all_items),
                    "total_comprobantes": len({i["codigo_comprobante"] for i in all_items}),
                },
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    print(f"Wrote {OUT_SQL} ({len(all_items)} ítems)")
    print(f"Wrote {OUT_JSON}")


if __name__ == "__main__":
    main()
